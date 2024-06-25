import sys
from openpyxl import load_workbook, Workbook

def insert_blank_rows(filename, N, M):
    wb = load_workbook(filename)
    sheet = wb.active

    new_wb = Workbook()
    new_sheet = new_wb.active

    for row in range(1, N + 1):
        for col in range(1, sheet.max_column + 1):
            new_sheet.cell(row=row, column=col, value=sheet.cell(row=row, column=col).value)

    for row in range(N + 1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            new_sheet.cell(row=row + M, column=col, value=sheet.cell(row=row, column=col).value)

    new_filename = "new_" + filename
    new_wb.save(new_filename)
    print(f"新しいファイルが作成されました: {new_filename}")

def main():
    if len(sys.argv) != 4:
        print("使用方法: python blankRowInserter.py <filename> <N> <M>")
        sys.exit(1)

    filename = sys.argv[1]
    try:
        N = int(sys.argv[2])
        M = int(sys.argv[3])
    except ValueError:
        print("NとMは整数である必要があります")
        sys.exit(1)

    insert_blank_rows(filename, N, M)

if __name__ == "__main__":
    main()

